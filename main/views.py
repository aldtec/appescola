from django.shortcuts import render
from .models import Calendario, Docente
# Usando no examplo
from django.http import HttpResponseRedirect
from .forms import PeriodoForm, NameForm

# Usado nos testes
from django.http import HttpResponse
import pendulum
pendulum.set_locale('pt-br')
from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl import load_workbook, Workbook
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Fill
from openpyxl.cell import Cell
from openpyxl.drawing.image import Image
import datetime
import pendulum
pendulum.set_locale('pt-br')
from django.conf.urls.static import static
from annoying.functions import get_object_or_None

# Para garantir acesso a imagens em produção/heroku
from django.contrib.staticfiles import finders

def come(request):

	def dsc(dia, mes, ano):
		date = pendulum.date(ano, mes, dia)
		mes = date.format('MMMM')
		ano = date.year
		fundo = ['dom','sáb']
		fmes = date.days_in_month
		x = date.format('ddd')
		if x in fundo:
			f = True
		else:
			f = False
		l = {
		"s"   : x[0].capitalize(),
		"sss" : x.capitalize(),
		"f"   : f,
		"fm"  : fmes,
		"mes" : mes,
		"ano" : ano
		}
		return l

	def dayoff(mes, ano):
		#pesqui = pendulum.parse(dtc)#parse(dtc.strftime("%Y-%m-%d"))
		#mes = pesqui.month
		#ano = pesqui.year

		# Filtra com mais de um argumento e ordena o resultado
		filtra = Calendario.objects.filter(data__year = ano).filter(data__month = mes).order_by('data')

		dici = []
		for dt in filtra:
			dici.append([
				dt.descricao.upper(),
				#dt.data.strftime("%a, %d, %m, %Y"),
				#pendulum.parse(dt.data.strftime("%Y-%m-%d")),
				dt.data.strftime("%Y-%m-%d"),
				#dt.data,
				dt.get_observ_display().upper()
				])

		return dici

	def dtlocalpx(mes, ano):
		data = pendulum.date(ano, mes, 1)
		fim = data.days_in_month
		x = pendulum.date(ano, mes, fim).add(days=1)
		ano = x.format('YYYY')
		mes = x.format('MMMM').capitalize()
		m = x.format('ddd')
		dia = x.day
		if m == 'dom':
			dia += 1
		elif m == 'sáb':
			dia += 2
			r = "São Paulo, "+str(dia)+" de "+mes+" de "+ano+"."
			return r

	def dtlocalhj(ano, mes):
		x = pendulum.date(ano, mes, 1)
		ano = x.format('YYYY')
		mes = x.format('MMMM').capitalize()
		dia = x.format('Do')
		l = {
			"s"   : "São Paulo, 1 de "+mes+" de "+ano+".",
			"pri" : "",
		"r"   : "São Paulo, "+dia+" de "+mes+" de "+ano+"."
		}
		return l

	resposta = '3,2020'
	res = resposta.split(',')
	mes = int(res[0])
	ano = int(res[1])
	ext = dsc(1, mes, ano)
	po = "excel/Ponto_docente.xlsx"
	wb = load_workbook(filename=po)
	ponto = "modelo"
	ps = wb[ponto]

	# Capa do ponto - falta converter para função
	capa  = "capa"
	cap = wb[capa]
	cap['A11'].value = ext['mes'].upper()
	#cap.add_image(logo, "B2")

	# Abertura do ponto - falta converter para função
	abert = "abertura"
	abt = wb[abert]
	abt['A13'].value = dtlocalhj(ano, mes)['s'] # Local 1º dia do mes

	# Para obter quantidades de dias no mes
	fim_mes = ext['fm']

	#Onde a linha começa no excel
	numero = 11
	
	#Para setar a cor de fundo das celulas
	fundo_cinza = PatternFill(fill_type='solid', start_color='BFBFBF', end_color='BFBFBF')#PatternFill(fill_type=None, start_color='A7A7A7', end_color='A7A7A7')

	ps['S4'].value = str(ext['mes'].capitalize())+" "+str(ext['ano'])

	#treta = dici[0][1].format('ddd')

	i = 1
	while i <= fim_mes:
		linha = 10 + i
		ps['A'+str(linha)].value = i
		week = dsc(i, mes, ano)
		diasfolga = dayoff(mes, ano)
		ps['B'+str(linha)].value = week['sss']
		total_colunas = 20
		init_colunas = 3
		connn = str(pendulum.date(ano, mes, i)) #str(ano)+"-"+str(mes)+"-"+str(i) #ano+"-"+mes+"-"+i

		#pendulum.parse(dt.data.strftime("%Y-%m-%d"))
		elem =  get_object_or_None(Calendario, data=pendulum.date(ano, mes, i))
		if elem:
			#mude o valor da coluna caso necessario
			ps.cell(row=linha, column=19, value=elem.get_observ_display().upper())
			while init_colunas <= total_colunas: 	
				ps.cell(row=linha, column=init_colunas).fill = fundo_cinza
				init_colunas += 1

		init_colunas = 3
		if week['f']: #is True:
			while init_colunas <= total_colunas: 	
				ps.cell(row=linha, column=init_colunas).fill = fundo_cinza
				init_colunas += 1
		i = int(i)
		i += 1


	testando = Docente.objects.all().order_by("rf_vinc")
	dicionario = []
	for cada in testando:
		teacher = Docente.objects.get(nome=cada)
		dicionario.append([
			teacher.nome.upper(), 
			teacher.rf_vinc, 
			teacher.qpe, 
			teacher.cargo,
			teacher.regencia,
			teacher.hor_col,
			teacher.turma,
			teacher.horario,
			teacher.get_jornada_display()
			])

	for cdprof in dicionario:
		tgnome = cdprof[1].replace('/', '-')
		target = wb.copy_worksheet(ps)
		result = finders.find('image/logo_pequeno.png')
		#logo = Image("static/image/logo_pequeno.png")
		logo = Image(result)
		logo.height = 70
		logo.width = 70
		target.add_image(logo, "B1")
		target.title = tgnome
		tg = wb[tgnome]
		tg['C6'].value = cdprof[0]
		tg['H6'].value = cdprof[1]
		tg['H7'].value = cdprof[2]
		tg['C7'].value = cdprof[3]
		tg['C8'].value = cdprof[4]
		tg['I8'].value = cdprof[5]
		tg['T8'].value = cdprof[6]
		tg['C9'].value = cdprof[7]
		tg['S7'].value = cdprof[8]


	final = "final"
	fl = wb[final]
	fl['A11'].value = dtlocalpx(mes, ano)
	target = wb.copy_worksheet(fl)
	target.title = 'Encerramento'

	wb.remove(wb.get_sheet_by_name(ponto))
	wb.remove(wb.get_sheet_by_name(final))

	wb.save("excel/ponto.xlsx")

	return HttpResponse("Deu certo!!!")

def get_name(request):
	if request.method == "POST":
		a = request.POST['resultado']
		po = "excel/Ponto_docente.xlsx"
		wb = load_workbook(filename=po)
		nome = a + ".xlsx"
		wb.save("excel/" + nome)
		return HttpResponse(a)
	else:
		lista = []
		today = pendulum.today('America/Sao_Paulo')
		mes = today.month
		ano = today.year
		#vai = str(mes)+","+str(ano)
		today = pendulum.today('America/Sao_Paulo').subtract(months=1)
		vai = str(today.month)+","+str(today.year)
		foi = today.format('MMMM').capitalize()+" - "+str(ano)
		#date = pendulum.date(ano, mes, dia)
		mes_passado = pendulum.today('America/Sao_Paulo').subtract(months=1)
		lista.append([vai, foi])
		x = 0
		while x <= 2:
			today = pendulum.today('America/Sao_Paulo').add(months=x)
			lista.append([
		 		str(today.month)+","+str(today.year), #vai
				today.format('MMMM').capitalize()+" - "+str(today.year) #foi
			])
			x += 1
		context = {
			
			"vum" : lista[0][0],
			"tum" : lista[0][1],
			"vdois": lista[1][0],
			"tdois": lista[1][1],
			"vtres": lista[2][0],
			"ttres": lista[2][1],
			"vquatro": lista[3][0],
			"tquatro": lista[3][1]
		}
		#vum = lista[0][0]
		#tum = lista[0][1]

		return render(request, "name.html", context)

def comee(request):
	listagem = []
	today = pendulum.today('America/Sao_Paulo')
	mes = today.month
	ano = today.year
	#vai = str(mes)+","+str(ano)
	today = pendulum.today('America/Sao_Paulo').subtract(months=1)
	vai = str(today.month)+","+str(today.year)
	foi = today.format('MMMM').capitalize()+" - "+str(ano)
	#date = pendulum.date(ano, mes, dia)
	mes_passado = pendulum.today('America/Sao_Paulo').add(months=12)
	listagem.append([vai, foi])
	x = 0
	while x <= 2:
		today = pendulum.today('America/Sao_Paulo').add(months=x)
		listagem.append([
	 		str(today.month)+","+str(today.year), #vai
			today.format('MMMM').capitalize()+" - "+str(today.year) #foi
		])
		x += 1
	return HttpResponse(mes_passado)

